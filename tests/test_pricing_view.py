"""The new engine shown in Price Pulse, alongside the existing suggestion.

The bridge in app/pricing_view.py is the single place that joins the pricing
engine to a database row, so every screen shows the same recommendation instead
of each assembling its own and quietly diverging.
"""

from __future__ import annotations

import re
import sys
from decimal import Decimal
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from fastapi.testclient import TestClient  # noqa: E402
from test_import_comparison_collection import _comparison_db, _product_id  # noqa: E402

from app.pricing_view import recommendation_for_row  # noqa: E402
from app.web.app import create_app  # noqa: E402


def test_a_row_produces_a_recommendation_with_its_reasoning() -> None:
    """A recommendation nobody can follow will not be trusted, so the parts
    that produced it have to travel with it."""
    row = {
        "our_current_price": "163.99",
        "current_cost": "107.19",
        "product_name": "DRIVE BELT",
        "manufacturer": "Polaris",
        "units_sold_12m": 8073,
        "sales_period": "12_months",
        "partzilla_selling_price": "189.99",
        "motosport_selling_price": "192.50",
        "chaparral_selling_price": "188.00",
    }

    result = recommendation_for_row(row, minimum_margin=Decimal("20"))

    assert result is not None
    assert result["action"] == "INCREASE"
    assert result["category"] == "Drivetrain / Transmission"
    assert result["sensitivity"] == "HIGH"
    assert result["sensitivity_factors"]
    assert result["reason"]
    assert result["competitor_confidence"] == "HIGH"


def test_the_sales_period_is_applied_before_scoring() -> None:
    """A quantity means nothing without the period behind it. The same number
    over six months is twice the demand, and must score accordingly."""
    base = {
        "our_current_price": "50.00",
        "current_cost": "25.00",
        "product_name": "DRIVE BELT",
        "units_sold_12m": 400,
        "partzilla_selling_price": "52.00",
    }

    annual = recommendation_for_row({**base, "sales_period": "12_months"})
    half_year = recommendation_for_row({**base, "sales_period": "6_months"})

    assert annual["annualized_qty"] == 400
    assert half_year["annualized_qty"] == 800
    assert half_year["sensitivity_score"] > annual["sensitivity_score"]
    assert "scaled up" in half_year["sales_period_note"]


def test_a_part_with_no_price_yields_nothing_rather_than_a_guess() -> None:
    assert recommendation_for_row({"our_current_price": "", "product_name": "DRIVE BELT"}) is None
    assert recommendation_for_row({"our_current_price": "0", "product_name": "DRIVE BELT"}) is None


def test_rejected_competitor_prices_are_reported_not_hidden() -> None:
    """Someone looking at a recommendation should be able to see that a
    competitor was ignored, and why, rather than wondering where it went."""
    result = recommendation_for_row(
        {
            "our_current_price": "177.66",
            "current_cost": "100.00",
            "product_name": "BATTERY GYZ20HA",
            "units_sold_12m": 120,
            "partzilla_selling_price": "177.66",
            "motosport_selling_price": "190.35",
            "chaparral_selling_price": "0.00",
        }
    )

    assert any("zero" in item for item in result["rejected_quotes"])


def test_an_out_of_stock_competitor_is_excluded_with_a_reason() -> None:
    result = recommendation_for_row(
        {
            "our_current_price": "100.00",
            "current_cost": "50.00",
            "product_name": "DRIVE BELT",
            "partzilla_selling_price": "80.00",
            "partzilla_availability_status": "out_of_stock",
            "motosport_selling_price": "105.00",
        }
    )

    assert any("out of stock" in item for item in result["rejected_quotes"])
    assert result["lowest_valid"] == "105.00"


def test_the_product_page_shows_the_recommendation_and_its_reasoning() -> None:
    database = _comparison_db("pricing_view_page.db")
    product_id = _product_id(database, "K-PRICE")

    page = TestClient(create_app(database), raise_server_exceptions=False).get(f"/products/{product_id}").text
    section = re.search(r"<h2>Suggested Pricing</h2>.*?</article>", page, re.S)

    assert section, "the product page should show the suggested pricing card"
    text = re.sub(r"<[^>]+>", " ", section.group(0))

    assert "Type of part" in text
    assert "How price sensitive" in text
    assert "Competitor prices used" in text
    # It must be clear this is advisory, since nothing is applied automatically.
    assert "Nothing is changed until you choose" in text


def test_the_product_page_still_loads_if_a_recommendation_cannot_be_made() -> None:
    """The recommendation is useful context, not the reason the page exists."""
    from app.web.queries import product_detail

    database = _comparison_db("pricing_view_missing.db")
    product_id = _product_id(database, "K-PRICE")

    detail = product_detail(database, product_id)

    assert detail is not None
    assert "pricing_recommendation" in detail


def test_the_margin_floor_comes_from_the_existing_rule() -> None:
    """One setting for both engines, so changing it in the rules screen changes
    both rather than the two drifting apart."""
    from app.pricing_view import minimum_margin_pct

    database = _comparison_db("pricing_view_margin.db")

    assert minimum_margin_pct(database) == Decimal("20")


def test_the_review_export_carries_the_new_engines_recommendation() -> None:
    """The export drives the actual pricing review, so a recommendation that
    only appears on a product page cannot be acted on across a thousand parts.
    The existing Suggested_Price is unchanged and the new columns sit beside
    it, so both can be compared in one place."""
    import tempfile
    from decimal import Decimal as D

    import openpyxl

    from app.comparison import ComparisonFilters, comparison_rows
    from app.exports.review_export import export_review

    database = _comparison_db("export_recommendation.db")
    rows = comparison_rows(database, ComparisonFilters())

    with tempfile.TemporaryDirectory() as tmp:
        path = export_review(rows, Path(tmp), minimum_margin=D("20"))
        sheet = openpyxl.load_workbook(path)["Pricing Review"]
        exported = list(sheet.iter_rows(values_only=True))

    header = exported[0]
    for column in (
        "Type_Of_Part",
        "Category_Confidence",
        "Qty_Sold_Annualized",
        "Sensitivity",
        "Sensitivity_Score",
        "New_Action",
        "New_Suggested_Price",
        "Competitor_Confidence",
        "Annual_Competitive_Price_Exposure",
        "Target_Percent_Of_Lowest",
        "Pricing_Rule_Applied",
        "Why",
    ):
        assert column in header, column

    # The old suggestion must still be there: both are shown, not replaced.
    assert "Suggested_Price" in header

    # And the new columns must actually be populated, not merely present.
    action_index = header.index("New_Action")
    why_index = header.index("Why")
    assert any(row[action_index] for row in exported[1:])
    assert any(row[why_index] for row in exported[1:])


def test_a_part_that_cannot_be_assessed_does_not_lose_the_export() -> None:
    """One unusual part must not cost someone the whole spreadsheet."""
    import tempfile
    from decimal import Decimal as D

    from app.exports.review_export import export_review

    with tempfile.TemporaryDirectory() as tmp:
        path = export_review(
            [{"internal_sku": "X", "manufacturer": "Polaris", "oem_part_number": "P1", "our_current_price": ""}],
            Path(tmp),
            minimum_margin=D("20"),
        )
        # Checked inside the context: the directory is removed on exit, so
        # testing afterwards would always fail regardless of the export.
        assert path.exists()

        import openpyxl

        sheet = openpyxl.load_workbook(path)["Pricing Review"]
        exported = list(sheet.iter_rows(values_only=True))

    # The row is still present, just without a recommendation.
    assert len(exported) == 2
    assert exported[1][2] == "P1"
